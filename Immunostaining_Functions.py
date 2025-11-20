def rename_and_copy_czi_files(main_folder_path):
    import os
    import shutil
    """
    Rename all .czi files in each subfolder of a directory to the name of the folder with an incremental number at the end,
    and copy them to a new "All" folder.

    Parameters
    ----------
    main_folder_path : str
        Path to the main folder containing subfolders with .czi files.
    Returns
    -------
    None
    """
    # First, rename all .czi files
    for root, dirs, files in os.walk(main_folder_path):
        for folder_name in dirs:
            if folder_name == "All":
                continue  # Skip the "All" folder
            folder_path = os.path.join(root, folder_name)
            
            # Get a list of all CZI files in the folder
            czi_files = [file for file in os.listdir(folder_path) if file.endswith(".czi")]
            
            # Sort the files to ensure they are renamed in the order they appear in the folder
            czi_files.sort()

            # Rename the CZI files
            for j, czi_file in enumerate(czi_files):
                new_name = f"{folder_name}_{j+1:02d}.czi"
                os.rename(os.path.join(folder_path, czi_file), os.path.join(folder_path, new_name))

    # Create the "All" folder if it doesn't exist
    all_folder_path = os.path.join(main_folder_path, "All")
    if not os.path.exists(all_folder_path):
        os.makedirs(all_folder_path)

    # Then, copy all renamed .czi files into the "All" folder
    for root, dirs, files in os.walk(main_folder_path):
        for folder_name in dirs:
            if folder_name == "All":
                continue  # Skip the "All" folder
            folder_path = os.path.join(root, folder_name)
            
            # Get a list of all renamed CZI files in the folder
            czi_files = [file for file in os.listdir(folder_path) if file.endswith(".czi")]
            
            # Copy the files to the "All" folder, skipping repeat files
            for file_name in czi_files:
                src_file_path = os.path.join(folder_path, file_name)
                dst_file_path = os.path.join(all_folder_path, file_name)
                
                if not os.path.exists(dst_file_path):
                    shutil.copy2(src_file_path, dst_file_path)

def Convert_czi_to_tiff(folder_path):
    import os
    import czifile
    from tifffile import imsave
    import shutil
    import cv2
    import numpy as np
    
    # Create the "CZI" folder in each folder path
    czi_folder_path = os.path.join(folder_path, "CZI")
    os.makedirs(czi_folder_path, exist_ok=True)

    for filename in os.listdir(folder_path):
        if filename.endswith(".czi"):
            czi_path = os.path.join(folder_path, filename)
            with czifile.CziFile(czi_path) as czi:
                image_arrays = czi.asarray()
                for channel_idx, channel_image in enumerate(image_arrays):
                    # Convert the image to 16-bit
                    channel_image_16bit = cv2.normalize(channel_image, None, 0, 65535, cv2.NORM_MINMAX, dtype=cv2.CV_16U)
                    
                    tiff_path = os.path.splitext(czi_path)[0] + f"_C_{channel_idx}.tiff"
                    imsave(tiff_path, channel_image_16bit)

            # Move the CZI files to the "CZI" folder
            new_czi_path = os.path.join(czi_folder_path, filename)
            shutil.move(czi_path, new_czi_path)

def Background_Subtraction(folder_path, radius, display, save):
    import os
    import warnings
    from skimage import io, img_as_float32
    from skimage.restoration import rolling_ball
    import numpy as np
    import cv2
    import matplotlib.pyplot as plt
    """
    Apply a 50-pixel radius rolling ball background subtraction to TIFF files in a folder, convert to 32-bit, and save the results.

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the TIFF images.
    radius : int, optional
        Radius of the rolling ball. Default is 50.

    Returns
    -------
    None
    """
    # Create a subfolder for the processed images
    processed_folder = os.path.join(folder_path, "0_Background Subtraction")
    os.makedirs(processed_folder, exist_ok=True)

    # Get a list of all TIFF files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith('.tiff')]

    if not image_files:
        return

    # Suppress low contrast image warnings
    warnings.filterwarnings("ignore", category=UserWarning, message=".*is a low contrast image.*")

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image
        image = io.imread(image_path)

        # Apply the rolling ball algorithm for background subtraction
        background = rolling_ball(image, radius=radius)
        subtracted_image = image - background

        # Ensure the pixel values are within the valid range
        subtracted_image = np.clip(subtracted_image, 0, np.max(subtracted_image))

        # Convert the image to 32-bit
        subtracted_image_32bit = img_as_float32(subtracted_image)

        if save:
            # Save the processed image to the processed folder
            processed_image_path = os.path.join(processed_folder, image_file)
            io.imsave(processed_image_path, subtracted_image_32bit)
        
        if display:
            # Display the original and background subtracted images side by side
            # fig, axes = plt.subplots(1, 2, figsize=(10, 5), dpi=DEFAULT_DPI)
            fig, axes = plt.subplots(1, 2, figsize=(10, 5), dpi=100)
            axes[0].imshow(image, cmap='gray')
            axes[0].set_title('Original Image')
            axes[0].axis('off')
            axes[1].imshow(subtracted_image_32bit, cmap='gray')
            axes[1].set_title('Background Subtracted Image')
            axes[1].axis('off')
            plt.show()


def optimize_contrast(folder_path, display, save):
    import skimage
    from skimage import io, exposure
    import matplotlib.pyplot as plt
    import os
    import numpy as np

    """Optimize the contrast of images in a folder and display the results

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the images

    Returns
    -------
    None
    """
    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith('.tiff')]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        img = skimage.io.imread(image_path)

        # Normalize the image to be between -1 and 1
        img = img / np.max(np.abs(img))

        # Optimize the contrast of the image
        img_optimized = exposure.equalize_adapthist(img)

        # Specify the file path of the optimized image
        optimized_image_path = os.path.join(folder_path, image_file)

        if save:
            # Save the optimized image
            skimage.io.imsave(optimized_image_path, img_optimized)

        if display:
            # Display the original image and the optimized image side by side
            fig, axes = plt.subplots(1, 2, figsize=(10, 5))
            axes[0].imshow(img, cmap='gray')
            axes[0].set_title('Original Image')
            axes[0].axis('off')
            axes[1].imshow(img_optimized, cmap='gray')
            axes[1].set_title('Optimized Image')
            axes[1].axis('off')
            plt.show()



def Colorize_Composite(folder_path, display_results=False, save_results=True):
    """
    Colorize specified TIFF images and create composite images.

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the TIFF images.
    display_results : bool, optional
        Whether to display the images. Default is False.
    save_results : bool, optional
        Whether to save the images. Default is True.

    Returns
    -------
    None
    """
    import os
    import glob
    import numpy as np
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    from skimage.io import imread as load_image
    from skimage.color import gray2rgb

    # Create a new output path that includes the "Colorized" folder
    output_path = os.path.join(folder_path, "0_Colorized")

    # Create the "Colorized" folder if it doesn't exist
    os.makedirs(output_path, exist_ok=True)

    def create_colormap(color):
        colors = ["black", color]
        return mcolors.LinearSegmentedColormap.from_list("", colors)
    
    # Get a list of base file names for all TIFF images ending in _C_0
    base_files = [os.path.splitext(os.path.basename(f))[0].rsplit('_C_0', 1)[0] for f in glob.glob(folder_path + "/*_C_0.tiff")]

    # Define the colors for each channel
    channel_colors = ["yellow", "green", "red", "blue"]
    rgb_colors = [(1, 1, 0), (0, 1, 0), (1, 0, 0), (0, 0, 1)]  # yellow, green, red, blue

    # Iterate through each base file name
    for base_name in base_files:
        channel_images = []
        for i in range(4):  # Include the fourth channel
            channel_path = os.path.join(folder_path, f"{base_name}_C_{i}.tiff")
            if os.path.exists(channel_path):
                channel_image = load_image(channel_path).astype(np.float32)  # Use float32 for compatibility
                channel_images.append(channel_image)
            else:
                print(f"Channel {i} not found for {base_name}")
                break

        if len(channel_images) != 4:
            print(f"Skipping {base_name} due to missing channels")
            continue

        cmap_colors = [create_colormap(color) for color in channel_colors]
        output_paths = []  # Create an empty list to store the output paths

        for i, channel_image in enumerate(channel_images):
            vmax = np.percentile(channel_image, 99.9)
            vmin = np.percentile(channel_image, 0.1)

            # Specify the output path with the original TIFF file name and channel number
            channel_output_path = os.path.join(output_path, f'{base_name}_C_{i}.png')
            
            plt.imshow(channel_image, cmap=cmap_colors[i], vmin=vmin, vmax=vmax)
            plt.axis(False)
          
            if save_results:
                # Save the image with the specified output path
                plt.savefig(channel_output_path, dpi=300, transparent=True, bbox_inches='tight', pad_inches=0)
                output_paths.append(channel_output_path)
            if display_results:
                plt.show()
            plt.close() 

        # Create Composite Image
        merged = np.zeros_like(gray2rgb(channel_images[0]), dtype=np.float32)
        for i in range(4):  # Include the fourth channel
            colored_image = gray2rgb(channel_images[i]) * np.array(rgb_colors[i], dtype=np.float32)
            merged += colored_image

        # Normalize the merged image to ensure values are within the valid range
        merged = merged / np.max(merged)

        fig = plt.figure(figsize=(merged.shape[1]/100, merged.shape[0]/100))
        plt.axis(False)
        plt.imshow(merged)

        # Save the composite image with the specified output path
        composite_output_path = os.path.join(output_path, f'{base_name}_composite.png')
        if save_results:
            plt.savefig(composite_output_path, dpi=300, transparent=True, bbox_inches='tight', pad_inches=0)
            output_paths.append(composite_output_path)
        if display_results:
            plt.show()
        plt.close()


def Apply_Scale_Bar(colorized_image_path, Objective, save_output, display_output):
    import os
    from PIL import Image
    import matplotlib.pyplot as plt
    from matplotlib_scalebar.scalebar import ScaleBar

    # Get a list of all TIFF and PNG files in the folder
    image_files = [file for file in os.listdir(colorized_image_path) if file.endswith('.tiff') or file.endswith('.png')]

    # Set scale bar parameters based on the Objective value
    if Objective == 20:
        scale = 0.335
        default_length_fraction = 0.22  
        composite_length_fraction = 0.13
    elif Objective == 10:
        scale = 0.641
        default_length_fraction = 0.11
        composite_length_fraction = 0.13
    else:
        raise ValueError("Unsupported Objective value. Only 10 and 20 are supported.")

    # Iterate over each image file
    for image_file in image_files:
        # Skip images ending in "_composite.png"
        if image_file.endswith("_composite.png"):
            continue

        # Construct the full path to the image file
        image_path = os.path.join(colorized_image_path, image_file)

        # Open the image file
        image = Image.open(image_path)

        # Determine the length fraction based on the file name
        length_fraction = default_length_fraction

        # Set the color of the scale bar
        Color_Shade = 'white'

        # Create a figure and axes
        fig, ax = plt.subplots()

        # Display the image
        ax.imshow(image, cmap='gray')

        # Add the scale bar
        scalebar = ScaleBar(scale, 'um', length_fraction=length_fraction, color=Color_Shade, box_color='None', location='lower right')

        # Add the scale bar to the axes
        ax.add_artist(scalebar)

        # Remove the axis labels and ticks
        ax.axis('off')

        if save_output:
            # Save the figure with the scale bar
            output_path = os.path.join(colorized_image_path, image_file)
            plt.savefig(output_path, dpi=300, transparent=True, bbox_inches='tight', pad_inches=0)
            plt.close(fig)

        if display_output:
            # Display the figure
            plt.show()

def Stack_Images_Row(colorized_image_path):
    import os
    import matplotlib.pyplot as plt
    import numpy as np
    from itertools import groupby
    from PIL import ImageFile
    ImageFile.LOAD_TRUNCATED_IMAGES = True

    # Create the "Compiled" folder if it does not exist
    output_path = os.path.join(colorized_image_path, "0_Stacked_Row")
    os.makedirs(output_path, exist_ok=True)

    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(colorized_image_path) if file.endswith('.png')]

    # Sort the image files based on their names
    image_files.sort()

    # Group the image files by their base names (before "_composite" or "_C")
    def get_base_name(filename):
        if "_composite" in filename:
            return filename.split("_composite")[0]
        elif "_C_" in filename:
            return filename.split("_C_")[0]
        else:
            return filename.rsplit('_', 1)[0]  # Fallback for other cases

    image_files_grouped = [list(group) for key, group in groupby(image_files, key=get_base_name)]

    # Iterate over the groups of image files
    for image_files in image_files_grouped:
        # Create a dictionary to store the images by type
        image_dict = {key: None for key in ['composite', '_C_0', '_C_1', '_C_2', '_C_3']}

        # Assign images to their respective slots in the dictionary
        for image_file in image_files:
            if image_file.endswith('composite.png'):
                image_dict['composite'] = image_file
            elif '_C_0.png' in image_file:
                image_dict['_C_0'] = image_file
            elif '_C_1.png' in image_file:
                image_dict['_C_1'] = image_file
            elif '_C_2.png' in image_file:
                image_dict['_C_2'] = image_file
            elif '_C_3.png' in image_file:
                image_dict['_C_3'] = image_file

        # Define the order of images to plot
        ordered_images = [
            image_dict['composite'],
            image_dict['_C_0'],
            image_dict['_C_1'],
            image_dict['_C_2'],
            image_dict['_C_3']
        ]

        # Extract the base name for the image set
        base_name = get_base_name(image_files[0])

        # Create a new figure for each group of images
        nrows = 1
        ncols = 5  # Always 5 images in a row
        fig, axes = plt.subplots(nrows, ncols, figsize=(5 * ncols, 10))

        # Make sure that axes is always a list
        axes = [axes] if ncols == 1 else axes

        # Iterate over the ordered images and plot them
        for i, image_file in enumerate(ordered_images):
            ax = axes[i]

            if image_file is not None:
                # Read the image
                image_path = os.path.join(colorized_image_path, image_file)
                image = plt.imread(image_path)
                # Set the title to the base name for the composite image, or the channel for others
                if i == 0:  # Composite image
                    ax.set_title(base_name, color='white', fontsize=8)
                else:  # Channel images
                    ax.set_title(image_file.split('_')[-1].split('.')[0], color='white', fontsize=8)
                ax.imshow(image)
            else:
                # Create a blank placeholder for missing images
                placeholder = np.ones((100, 100, 3))  # White blank image
                ax.imshow(placeholder)
                ax.set_title("Missing", color='red', fontsize=8)

            ax.axis('off')  # Remove the axis

        # Save the figure
        output_file = os.path.join(output_path, base_name + '.png')
        plt.tight_layout()
        plt.savefig(output_file, dpi=300, transparent=True, bbox_inches='tight', pad_inches=0)
        plt.close(fig)  # Close the figure to free memory

def Stack_Images_Column(colorized_image_path):
    import os
    import matplotlib.pyplot as plt
    import numpy as np
    from itertools import groupby
    from PIL import ImageFile
    ImageFile.LOAD_TRUNCATED_IMAGES = True

    # Create the "Compiled" folder if it does not exist
    output_path = os.path.join(colorized_image_path, "0_Stacked_Column")
    os.makedirs(output_path, exist_ok=True)

    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(colorized_image_path) if file.endswith('.png')]

    # Sort the image files based on their names
    image_files.sort()

    # Group the image files by their base names (before "_composite" or "_C")
    def get_base_name(filename):
        if "_composite" in filename:
            return filename.split("_composite")[0]
        elif "_C_" in filename:
            return filename.split("_C_")[0]
        else:
            return filename.rsplit('_', 1)[0]  # Fallback for other cases

    image_files_grouped = [list(group) for key, group in groupby(image_files, key=get_base_name)]

    # Iterate over the groups of image files
    for image_files in image_files_grouped:
        # Create a dictionary to store the images by type
        image_dict = {key: None for key in ['composite', '_C_0', '_C_1', '_C_2', '_C_3']}

        # Assign images to their respective slots in the dictionary
        for image_file in image_files:
            if image_file.endswith('composite.png'):
                image_dict['composite'] = image_file
            elif '_C_0.png' in image_file:
                image_dict['_C_0'] = image_file
            elif '_C_1.png' in image_file:
                image_dict['_C_1'] = image_file
            elif '_C_2.png' in image_file:
                image_dict['_C_2'] = image_file
            elif '_C_3.png' in image_file:
                image_dict['_C_3'] = image_file

        # Define the order of images to plot
        ordered_images = [
            image_dict['composite'],
            image_dict['_C_0'],
            image_dict['_C_1'],
            image_dict['_C_2'],
            image_dict['_C_3']
        ]

        # Extract the base name for the image set
        base_name = get_base_name(image_files[0])

        # Create a new figure for each group of images
        ncols = 1
        nrows = 5  # Always 5 images in a column
        fig, axes = plt.subplots(nrows, ncols, figsize=(10, 5 * nrows))

        # Make sure that axes is always a list
        axes = [axes] if nrows == 1 else axes

        # Iterate over the ordered images and plot them
        for i, image_file in enumerate(ordered_images):
            ax = axes[i]

            if image_file is not None:
                # Read the image
                image_path = os.path.join(colorized_image_path, image_file)
                image = plt.imread(image_path)
                ax.imshow(image)
            else:
                # Create a blank placeholder for missing images
                placeholder = np.ones((100, 100, 3))  # White blank image
                ax.imshow(placeholder)

            ax.axis('off')  # Remove the axis

        # Save the figure
        output_file = os.path.join(output_path, base_name + '.png')
        plt.tight_layout()
        plt.savefig(output_file, dpi=300, transparent=True, bbox_inches='tight', pad_inches=0)
        plt.close(fig)  # Close the figure to free memory


## QUANTIFICATION

def Threshold_Test(folder_path):
    import os
    from skimage import io
    from skimage.filters import try_all_threshold
    import matplotlib.pyplot as plt
    """Test threshold options for images in a folder

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the images

    Returns
    -------
    None
    """
    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith('.tiff') or file.endswith('.tif')]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image
        img = io.imread(image_path)
        fig, ax = try_all_threshold(img, figsize=(20, 18), verbose=False)
        plt.title(image_file)  # Add the image file name as the title
        plt.show()

def Apply_mean_threshold(folder_path):
    import os
    from skimage import io, filters
    """Apply mean threshold filter to images in a folder

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the images

    Returns
    -------
    None
    """
    # Create a subfolder for the thresholded images
    result_folder = os.path.join(folder_path, "Thresholded")
    os.makedirs(result_folder, exist_ok=True)

    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith(('.tiff', '.tif', '.png', '.jpg'))]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image
        img = io.imread(image_path)

        # Apply mean threshold filter
        threshold_value = filters.threshold_mean(img)
        binary_image = img > threshold_value

        # Save the thresholded image to the result folder
        result_image_path = os.path.join(result_folder, image_file)
        io.imsave(result_image_path, binary_image.astype('uint8') * 255)

def Apply_otsu_threshold(folder_path):
    import os
    from skimage import io, filters
    """Apply Otsu threshold filter to images in a folder

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the images

    Returns
    -------
    None
    """
    # Create a subfolder for the thresholded images
    result_folder = os.path.join(folder_path, "Thresholded")
    os.makedirs(result_folder, exist_ok=True)

    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith(('.tiff', '.tif', '.png', '.jpg'))]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image
        img = io.imread(image_path)

        # Apply Otsu threshold filter
        threshold_value = filters.threshold_otsu(img)
        binary_image = img > threshold_value

        # Save the thresholded image to the result folder
        result_image_path = os.path.join(result_folder, image_file)
        io.imsave(result_image_path, binary_image.astype('uint8') * 255)

def Apply_mean_threshold_Astrocyte_Neurons(folder_path): #Auto apply to only astrocytes and neurons
    import os
    from skimage import io, filters
    """Apply mean threshold filter to images in a folder

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the images

    Returns
    -------
    None
    """
    # Create a subfolder for the thresholded images
    result_folder = os.path.join(folder_path, "Thresholded")
    os.makedirs(result_folder, exist_ok=True)

    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith('_C_1.tiff') or file.endswith('_C_2.tiff')]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image
        img = io.imread(image_path)

        # Apply mean threshold filter
        threshold_value = filters.threshold_mean(img)
        binary_image = img > threshold_value

        # Save the thresholded image to the result folder
        result_image_path = os.path.join(result_folder, image_file)
        io.imsave(result_image_path, binary_image.astype('uint8') * 255)

def Apply_otsu_threshold_Nuclei_Microglia(folder_path): #Auto apply to only microglia and DAPI
    import os
    from skimage import io, filters
    """Apply Otsu threshold filter to images in a folder

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the images

    Returns
    -------
    None
    """
    # Create a subfolder for the thresholded images
    result_folder = os.path.join(folder_path, "Thresholded")
    os.makedirs(result_folder, exist_ok=True)

    # Get a list of all image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith(('.tiff', '.png', '.jpg'))]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image
        img = io.imread(image_path)

        # Apply Otsu threshold filter
        threshold_value = filters.threshold_otsu(img)
        binary_image = img > threshold_value

        # Save the thresholded image to the result folder
        result_image_path = os.path.join(result_folder, image_file)
        io.imsave(result_image_path, binary_image.astype('uint8') * 255)

def Remove_Islands(folder_path, Microglial_disk, Astrocytes_disk, Neurons_disk, display_results, save_results):
    from skimage.morphology import disk
    from skimage import morphology
    import os
    from skimage import io
    import matplotlib.pyplot as plt
    """Process all TIFF images in a folder by applying white tophat transformation and saving the final images.
    
    NO DAPI DISK

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the TIFF images
    Microglia_disk : int
        Disk radius for Microglia images (C_0)
    Astrocytes_disk : int
        Disk radius for Astrocytes images (C_1)
    Neurons_disk : int
        Disk radius for Neurons images (C_2)
    display_results : bool, optional
        Whether to display the images. Default is False.
    save_results : bool, optional
        Whether to save the images. Default is True.

    Returns
    -------
    None
    """
    # Get a list of all TIFF image files in the folder
    image_files = [file for file in os.listdir(folder_path) if file.endswith('.tiff')]

    # Define the disk radius for different channels
    disk_radius = {
        '_C_0.tiff': Microglial_disk,
        '_C_1.tiff': Astrocytes_disk,
        '_C_2.tiff': Neurons_disk
    }

    # Iterate over each image file
    for image_file in image_files:
        # Determine the appropriate disk radius based on the file name
        radius = None
        for key in disk_radius:
            if image_file.endswith(key):
                radius = disk_radius[key]
                break

        if radius is None:
            print(f"Skipping {image_file} due to unmatched channel")
            continue

        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image file into a NumPy array
        image = io.imread(image_path)

        footprint = disk(radius)
        res = morphology.white_tophat(image, footprint)

        processed_image = image - res

        if save_results:
            final_image_path = os.path.join(folder_path, image_file)
            io.imsave(final_image_path, processed_image)

        if display_results:
            plt.figure(figsize=(10, 10))
            plt.imshow(processed_image, cmap='gray')
            plt.title(image_file)
            plt.axis('off')
            plt.show()

def Remove_DAPI_Islands(folder_path, disk_radius, display_results, save_results):
    from skimage.morphology import disk
    from skimage import morphology
    import os
    from skimage import io
    import matplotlib.pyplot as plt
    """Process all TIFF images in a folder by applying white tophat transformation and saving the final images.

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the TIFF images
    disk_radius : int
        Disk radius for the white tophat transformation
    display_results : bool, optional
        Whether to display the images. Default is False.
    save_results : bool, optional
        Whether to save the images. Default is True.

    Returns
    -------
    None
    """
    # Get a list of all TIFF image files in the folder that end with _C_3.tiff
    image_files = [file for file in os.listdir(folder_path) if file.endswith('_C_3.tiff')]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the image file into a NumPy array
        image = io.imread(image_path)

        footprint = disk(disk_radius)
        res = morphology.white_tophat(image, footprint)

        processed_image = image - res

        if save_results:
            final_image_path = os.path.join(folder_path, image_file)
            io.imsave(final_image_path, processed_image)

        if display_results:
            plt.figure(figsize=(10, 10))
            plt.imshow(processed_image, cmap='gray')
            plt.title(image_file)
            plt.axis('off')
            plt.show()

def Watershed_Segmentation(folder_path, min_distance, display_results=False, save_results=True):
    from skimage.segmentation import watershed
    from skimage.feature import peak_local_max
    from scipy import ndimage as ndi
    import numpy as np
    import matplotlib.pyplot as plt
    import skimage.io as io
    import os
    """Perform watershed segmentation on all images in the input folder.

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the thresholded images.
    min_distance : int
        Minimum number of pixels separating peaks in a region of interest.
    display_results : bool, optional
        Whether to display the images. Default is False.
    save_results : bool, optional
        Whether to save the images. Default is True.

    Returns
    -------
    None
    """
    # Get a list of all TIFF image files in the folder that end with _C_3.tiff
    image_files = [file for file in os.listdir(folder_path) if file.endswith('_C_3.tiff')]

    # Iterate over each image file
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)

        # Read the input image
        image = io.imread(image_path)

        # Compute the Euclidean distance from the background
        distance = ndi.distance_transform_edt(image)

        # Find the local maxima in the distance map
        coordinates = peak_local_max(distance, labels=image, min_distance=min_distance)

        # Create an empty array of zeros with the same shape as the image
        markers = np.zeros_like(image, dtype=int)

        # Label the local maxima coordinates
        markers[coordinates[:, 0], coordinates[:, 1]] = np.arange(len(coordinates)) + 1

        # Perform watershed segmentation
        labels = watershed(-distance, markers, mask=image)

        if save_results:
            # Save the segmented image
            segmented_image_path = os.path.join(folder_path, f"segmented_{image_file}")
            io.imsave(segmented_image_path, labels.astype(np.uint8))

        if display_results:
            plt.figure(figsize=(10, 10))
            plt.imshow(labels, cmap='nipy_spectral')
            plt.title(f"Segmented {image_file}")
            plt.axis('off')
            plt.show()

def DAPI_Count(folder_path, min_area, max_area, display_results=False, save_results=True):
    import cv2
    import os
    import pandas as pd
    import matplotlib.pyplot as plt
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    """Count the number of contours in each image file in the specified folder that meet the contour area specification.
    
    Parameters
    ----------
    folder_path : str
        Path to the folder containing the image files.
    min_area : int
        Minimum contour area.
    max_area : int
        Maximum contour area.
    display_results : bool, optional
        Whether to display the images. Default is False.
    save_results : bool, optional
        Whether to save the images. Default is True.
    
    Returns
    -------
    None
    """
    if save_results:
        # Check if the Excel file already exists
        excel_file_path = os.path.join(folder_path, "DAPI_Count.xlsx")

        # Create a new folder to save the images with contour overlay
        contour_folder_path = os.path.join(folder_path, "DAPI Count Contours")
        os.makedirs(contour_folder_path, exist_ok=True)       

        if os.path.exists(excel_file_path):
            # Load the existing workbook
            workbook = load_workbook(excel_file_path)
        else:
            # Create a new workbook
            workbook = Workbook()

        # Get the active sheet
        sheet = workbook.active

        # Write the header row if the sheet is empty
        if sheet.dimensions == 'A1:A1':
            header_row = ["Image File", "DAPI Count"]
            sheet.append(header_row)

    # Read all image files in the folder that end with _C_3.tiff
    image_files = [file for file in os.listdir(folder_path) if file.endswith('_C_3.tiff')]
    
    for image_file in image_files:
        # Construct the full path to the image file
        image_path = os.path.join(folder_path, image_file)
            
        image = cv2.imread(image_path)

        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)   
        contours, _ = cv2.findContours(gray_image, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE) # Find contours in the grayscale image
            
        # Initialize a counter for the areas that meet the contour area specification
        count = 0
            
        # Iterate over the contours
        for contour in contours:
            # Calculate the area of the contour
            area = cv2.contourArea(contour)
                
            # Check if the area meets the contour area specification
            if min_area <= area <= max_area:
                count += 1
                # Draw the contour on the original image
                cv2.drawContours(image, [contour], -1, (0, 255, 0), 2)

        if save_results:
            # Save the plot result to the contour folder
            plot_file_path = os.path.join(contour_folder_path, image_file.split('.')[0] + '.png')
            cv2.imwrite(plot_file_path, image)
         
            # Write the DataFrame to the sheet
            data = [[image_file, count]]
            df = pd.DataFrame(data, columns=["Image File", "DAPI Count"])
            for row in dataframe_to_rows(df, index=False, header=False):
                sheet.append(row)

        if display_results:
            plt.figure(figsize=(10, 10))
            plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
            plt.title(image_file)
            plt.axis('off')
            plt.show()

    if save_results:
        # Save the workbook to the specified output file
        workbook.save(excel_file_path)


def Measure_Surface_Area_Coverage(folder_path, display_results=False, save_results=True):
    import cv2
    import numpy as np
    import os
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import load_workbook
    import matplotlib.pyplot as plt

    """
    Measure surface area coverage for C_0, C_1, and C_2 images and save the results to an Excel sheet.

    Parameters
    ----------
    folder_path : str
        Path to the folder containing the image files.
    display_results : bool, optional
        Whether to display the images. Default is False.
    save_results : bool, optional
        Whether to save the images. Default is True.

    Returns
    -------
    None
    """
    # Create the output file path
    output_file = os.path.join(folder_path, "Surface_Area_Coverage.xlsx")

    if save_results:
        # Check if the Excel file already exists
        if os.path.exists(output_file):
            # Delete the existing file
            os.remove(output_file)
        
        # Create a new workbook
        workbook = Workbook()
        workbook.save(output_file)
        
        # Create a new folder to save the contour overlay images
        contour_folder = os.path.join(folder_path, "Surface Area Coverage Contours")
        os.makedirs(contour_folder, exist_ok=True)

    # Initialize the dictionary to store white area percentages
    white_area_percentages = {}

    # Iterate over each file in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".tiff") or filename.endswith(".png"):
            # Construct the full path to the image file
            image_path = os.path.join(folder_path, filename)

            # Read the image as a grayscale image
            image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)

            # Threshold the image to obtain a binary mask
            _, binary_mask = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY)

            # Count the number of white pixels in the binary mask
            white_pixels = np.count_nonzero(binary_mask)

            # Calculate the total number of pixels in the image
            total_pixels = image.shape[0] * image.shape[1]

            # Calculate the percentage of white area
            white_area_percentage = (white_pixels / total_pixels) * 100

            # Determine the base name and channel type based on the filename
            base_name = filename.rsplit('_C_', 1)[0]
            if "_C_0" in filename:
                channel_type = "Microglia"
            elif "_C_1" in filename:
                channel_type = "Astrocyte"
            elif "_C_2" in filename:
                channel_type = "Neuron"
            else:
                continue  # Skip files that do not match the required channels

            # Append the data to the dictionary
            if base_name not in white_area_percentages:
                white_area_percentages[base_name] = {"Microglia": None, "Astrocyte": None, "Neuron": None}
            white_area_percentages[base_name][channel_type] = white_area_percentage

            # Plot the image with the white area outlined in purple
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
            contours, _ = cv2.findContours(binary_mask, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            cv2.drawContours(image, contours, -1, (128, 0, 128), 2)
            
            if display_results:
                plt.figure(figsize=(10, 10))
                plt.imshow(image)
                plt.title(f"{filename}")
                plt.axis('off')
                plt.show()
            
            if save_results:
                # Save the contour overlay image to the contour folder
                contour_image_path = os.path.join(contour_folder, f"{os.path.splitext(filename)[0]}.png")
                plt.imsave(contour_image_path, image)
                plt.close()

    if save_results:
        # Create a DataFrame from the white area percentages
        df = pd.DataFrame.from_dict(white_area_percentages, orient='index').reset_index()
        df.columns = ["Image File", "Microglia", "Astrocyte", "Neuron"]

        # Load the existing workbook
        workbook = load_workbook(output_file)

        # Select the active sheet
        sheet = workbook.active

        # Write the DataFrame to the sheet
        rows = dataframe_to_rows(df, index=False, header=True)
        
        # Only write headers if sheet is empty
        if sheet.max_row == 1:
            headers = ["Image File", "Microglia", "Astrocyte", "Neuron"]
            sheet.append(headers)
        
        for row in rows:
            sheet.append(row)

        # Save the workbook to the specified output file
        workbook.save(output_file)
